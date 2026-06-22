# -*- coding: utf-8 -*-
"""Chia file tổng hợp sản phẩm thành N shop:
 - Gom danh mục liên quan vào cùng shop (sort theo đường dẫn danh mục, chia khối liền kề)
 - Cân bằng số sản phẩm mỗi shop
 - Ít trùng tên sản phẩm trong cùng 1 shop nhất (interleave nhóm trùng)
Mỗi shop -> 1 sheet, tên shop1..shopN.
"""
import re
import sys
from collections import defaultdict, Counter
import openpyxl
from openpyxl import Workbook

SRC = r"C:\Users\Ng Xuan Mui\Downloads\tonghop-file-361shop-29207sp-2026-06-21_134334.xlsx"
OUT = r"C:\Users\Ng Xuan Mui\Downloads\chia-58shop-2026-06-21.xlsx"
N_SHOPS = 58

CAT_COL = 7   # "Danh mục"
NAME_COL = 5  # "Tên sp"
IMG_COL = 14  # "Ảnh"

def norm_name(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())

def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

    header = None
    by_cat = defaultdict(list)   # category path -> list of rows (each row = list of cells)
    total = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        it = ws.iter_rows(values_only=True)
        h = next(it, None)
        if h is None:
            continue
        if header is None:
            header = list(h)
        for r in it:
            if r is None:
                continue
            if all(x is None or str(x).strip() == "" for x in r):
                continue
            row = list(r)
            # đảm bảo đủ số cột
            while len(row) < len(header):
                row.append(None)
            cat = (str(row[CAT_COL]).strip() if len(row) > CAT_COL and row[CAT_COL] is not None else "")
            if not cat:
                cat = "(trống)"
            by_cat[cat].append(row)
            total += 1
    wb.close()

    print("Tổng sản phẩm:", total)
    print("Số danh mục (leaf):", len(by_cat))

    # Sort danh mục theo đường dẫn -> danh mục liên quan nằm cạnh nhau
    cats_sorted = sorted(by_cat.keys())

    # Xây dựng thứ tự sản phẩm toàn cục
    ordered = []
    for cat in cats_sorted:
        prods = by_cat[cat]
        # interleave theo nhóm trùng tên (tách bản trùng ra xa nhau)
        groups = defaultdict(list)
        for idx, p in enumerate(prods):
            key = norm_name(p[NAME_COL] if len(p) > NAME_COL else "")
            if not key:
                key = (str(p[IMG_COL]).strip() if len(p) > IMG_COL and p[IMG_COL] is not None else "")
            if not key:
                key = f"__uniq_{cat}_{idx}"
            groups[key].append(p)
        glists = list(groups.values())
        # nhóm trùng nhiều đi trước để được tách rộng nhất
        glists.sort(key=len, reverse=True)
        maxlen = max(len(g) for g in glists)
        for i in range(maxlen):
            for g in glists:
                if i < len(g):
                    ordered.append(g[i])

    assert len(ordered) == total

    # Gán shop theo vị trí -> cân bằng + khối liền kề (danh mục liên quan cùng shop)
    shops = [[] for _ in range(N_SHOPS)]
    for pos, p in enumerate(ordered):
        s = pos * N_SHOPS // total
        if s >= N_SHOPS:
            s = N_SHOPS - 1
        shops[s].append(p)

    # Ghi file ra
    out = Workbook()
    out.remove(out.active)
    dup_within_total = 0
    print("\nShop | SP | #DanhMục | Trùng-trong-shop | Danh mục chính")
    for i, rows in enumerate(shops, start=1):
        ws = out.create_sheet(title=f"shop{i}")
        ws.append(header)
        for row in rows:
            ws.append(row)
        # thống kê
        cat_counter = Counter()
        name_counter = Counter()
        for row in rows:
            c = (str(row[CAT_COL]).strip() if row[CAT_COL] is not None else "(trống)")
            cat_counter[c] += 1
            nm = norm_name(row[NAME_COL])
            if nm:
                name_counter[nm] += 1
        dup_within = sum(v - 1 for v in name_counter.values() if v > 1)
        dup_within_total += dup_within
        top_cat = cat_counter.most_common(1)[0][0] if cat_counter else ""
        # rút gọn tên danh mục chính
        tc = top_cat if len(top_cat) <= 50 else top_cat[:47] + "..."
        print(f"shop{i:<2} | {len(rows):4d} | {len(cat_counter):3d}      | {dup_within:3d}              | {tc}")

    out.save(OUT)
    print("\nĐã lưu:", OUT)
    print("Tổng số sản phẩm trùng tên còn chung shop (dư):", dup_within_total,
          f"  (trên tổng {total})")
    counts = [len(s) for s in shops]
    print(f"Cỡ shop: min={min(counts)} max={max(counts)} (lệch {max(counts)-min(counts)})")

if __name__ == "__main__":
    main()
