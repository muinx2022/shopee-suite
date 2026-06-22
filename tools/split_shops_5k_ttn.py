# -*- coding: utf-8 -*-
"""CHỈ lấy ngành 'Thời Trang Nữ', loại hết ngành khác, rồi chia 58 shop x 5000 (nhân bản).
Giống split_shops_5k.py nhưng có bộ lọc danh mục cấp 1.
"""
import re
import random
import heapq
from collections import Counter
import openpyxl
from openpyxl import Workbook

random.seed(20260621)

SRC = r"C:\Users\Ng Xuan Mui\Downloads\tonghop-file-361shop-29207sp-2026-06-21_134334.xlsx"
OUT = r"C:\Users\Ng Xuan Mui\Downloads\chia-58shop-5000sp-ThoiTrangNu-2026-06-21.xlsx"
N_SHOPS = 58
PER = 5000
FILTER_L1 = "thời trang nữ"   # chỉ giữ danh mục cấp 1 này

CAT_COL, NAME_COL, IMG_COL, ITEM_COL = 7, 5, 14, 16

def norm_name(s):
    return re.sub(r"\s+", " ", str(s).strip().lower()) if s is not None else ""

def lvl1(cat):
    return (str(cat).split(">")[0].strip().lower()) if cat else ""

def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    header = None
    all_rows = []
    kept_other = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        it = ws.iter_rows(values_only=True)
        h = next(it, None)
        if h is None:
            continue
        if header is None:
            header = list(h)
        for r in it:
            if r is None or all(x is None or str(x).strip() == "" for x in r):
                continue
            row = list(r)
            while len(row) < len(header):
                row.append(None)
            if lvl1(row[CAT_COL]) == FILTER_L1:      # LỌC: chỉ Thời Trang Nữ
                all_rows.append(row)
            else:
                kept_other += 1
    wb.close()
    print(f"Listing 'Thời Trang Nữ': {len(all_rows)}   (đã loại {kept_other} sp ngành khác)")

    # Gộp trùng tên -> giữ bản đầu tiên
    seen = {}
    for r in all_rows:
        key = norm_name(r[NAME_COL])
        if not key:
            key = "img:" + (str(r[IMG_COL]).strip() if r[IMG_COL] is not None else "")
        if not key or key == "img:":
            key = "id:" + (str(r[ITEM_COL]).strip() if r[ITEM_COL] is not None else str(len(seen)))
        if key not in seen:
            seen[key] = r
    products = list(seen.values())
    products.sort(key=lambda r: (str(r[CAT_COL] or ""), norm_name(r[NAME_COL])))
    M = len(products)
    TOTAL = N_SHOPS * PER
    print(f"SP distinct (sau gộp trùng tên): {M}")
    print(f"Tổng ô: {N_SHOPS} x {PER} = {TOTAL}  -> mỗi sp ~{TOTAL/M:.1f} shop")

    # Bậc mỗi sp
    base = TOTAL // M
    extra = TOTAL - base * M
    deg = [base] * M
    for n in range(extra):
        deg[(n * M) // extra] += 1

    # Phân bổ cân bằng theo dung lượng còn lại (tie-break ngẫu nhiên)
    cap = [PER] * N_SHOPS
    shops = [[] for _ in range(N_SHOPS)]
    order = list(range(M))
    random.shuffle(order)
    for i in order:
        d = deg[i]
        chosen = heapq.nlargest(d, range(N_SHOPS), key=lambda s: cap[s] + random.random())
        for s in chosen:
            cap[s] -= 1
            shops[s].append(products[i])
    assert all(c == 0 for c in cap), f"cap dư: {[c for c in cap if c]}"

    # Kiểm tra
    use = Counter()
    name_dup_in_shop = 0
    sets = []
    for s in shops:
        assert len(s) == PER
        items = [r[ITEM_COL] for r in s]
        assert len(set(items)) == PER, "có listing lặp trong shop!"
        name_dup_in_shop += sum(v - 1 for v in Counter(norm_name(r[NAME_COL]) for r in s).values() if v > 1)
        for r in s:
            use[r[ITEM_COL]] += 1
        sets.append(set(items))
    uc = Counter(use.values())
    ov_min, ov_max, ov_sum, npair = 10**9, 0, 0, 0
    for i in range(N_SHOPS):
        for j in range(i + 1, N_SHOPS):
            o = len(sets[i] & sets[j]); ov_min = min(ov_min, o); ov_max = max(ov_max, o); ov_sum += o; npair += 1
    print("\n--- KIỂM TRA ---")
    print("Mỗi shop:", PER, "sp khác nhau | Lặp trong shop: 0")
    print("Cặp listing trùng tên trong 1 shop:", name_dup_in_shop)
    print("Số lần dùng mỗi sp:", dict(sorted(uc.items())))
    print(f"Overlap 2 shop: min={ov_min} avg={ov_sum/npair:.0f} max={ov_max} (trên {PER})")

    # Ghi file
    out = Workbook(write_only=True)
    for i in range(1, N_SHOPS + 1):
        ws = out.create_sheet(title=f"shop{i}")
        rows = shops[i - 1]
        rows.sort(key=lambda r: (str(r[CAT_COL] or ""), norm_name(r[NAME_COL])))
        ws.append(header)
        for r in rows:
            ws.append(r)
    out.save(OUT)
    print("\nĐã lưu:", OUT)

if __name__ == "__main__":
    main()
