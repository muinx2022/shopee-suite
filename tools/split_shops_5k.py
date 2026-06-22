# -*- coding: utf-8 -*-
"""Chia kho sản phẩm thành N shop, MỖI shop đúng PER sản phẩm (nhân bản).
 - Gộp listing trùng tên -> 1 sp đại diện (không shop nào có 2 listing trùng tên)
 - Round-robin bước N: mỗi sp xuất hiện ở ~ceil(TOTAL/M) shop KHÁC NHAU, rải đều
 - Mỗi shop đúng PER sp khác nhau; 2 shop bất kỳ trùng nhau ít & đồng đều
 - Mỗi sheet sắp theo danh mục (danh mục liên quan nằm gần nhau)
"""
import re
import random
import heapq
from collections import defaultdict, Counter
import openpyxl
from openpyxl import Workbook

random.seed(20260621)

SRC = r"C:\Users\Ng Xuan Mui\Downloads\tonghop-file-361shop-29207sp-2026-06-21_134334.xlsx"
OUT = r"C:\Users\Ng Xuan Mui\Downloads\chia-58shop-5000sp-2026-06-21.xlsx"
N_SHOPS = 58
PER = 5000

CAT_COL, NAME_COL, IMG_COL, ITEM_COL = 7, 5, 14, 16

def norm_name(s):
    return re.sub(r"\s+", " ", str(s).strip().lower()) if s is not None else ""

def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    header = None
    all_rows = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        it = ws.iter_rows(values_only=True)
        h = next(it, None)
        if h is None:
            continue
        if header is None:
            header = list(h)
        for r in it:
            if r is None or all(x is None or str(x).strip() == "" for x in r):
                continue
            row = list(r)
            while len(row) < len(header):
                row.append(None)
            all_rows.append(row)
    wb.close()
    print("Listing gốc:", len(all_rows))

    # Gộp trùng tên -> giữ bản đầu tiên làm đại diện
    seen = {}
    for r in all_rows:
        key = norm_name(r[NAME_COL])
        if not key:
            key = "img:" + (str(r[IMG_COL]).strip() if r[IMG_COL] is not None else "")
        if not key or key == "img:":
            key = "id:" + (str(r[ITEM_COL]).strip() if r[ITEM_COL] is not None else str(len(seen)))
        if key not in seen:
            seen[key] = r
    products = list(seen.values())
    # Sắp theo danh mục -> tên (danh mục liên quan cạnh nhau)
    products.sort(key=lambda r: (str(r[CAT_COL] or ""), norm_name(r[NAME_COL])))
    M = len(products)
    TOTAL = N_SHOPS * PER
    print(f"SP distinct (sau gộp trùng tên): {M}")
    print(f"Tổng ô: {N_SHOPS} shop x {PER} = {TOTAL}")
    print(f"Mỗi sp xuất hiện: {TOTAL//M}-{(TOTAL+M-1)//M} lần")

    # Bậc (số shop) mỗi sp: rải đều phần "11 lần"
    base = TOTAL // M           # mỗi sp ít nhất 'base' shop
    extra = TOTAL - base * M    # số sp được +1 shop
    deg = [base] * M
    for n in range(extra):
        deg[(n * M) // extra] += 1   # rải đều các sp bậc cao

    # Phân bổ cân bằng: mỗi sp chọn deg[i] shop có DUNG LƯỢNG còn lại lớn nhất
    # (tie-break ngẫu nhiên) -> đếm mỗi shop tiến tới đúng PER, overlap đồng đều.
    cap = [PER] * N_SHOPS
    shops = [[] for _ in range(N_SHOPS)]
    order = list(range(M))
    random.shuffle(order)
    for i in order:
        d = deg[i]
        # d shop có cap lớn nhất, tie-break ngẫu nhiên
        chosen = heapq.nlargest(d, range(N_SHOPS), key=lambda s: cap[s] + random.random())
        for s in chosen:
            cap[s] -= 1
            shops[s].append(products[i])
    assert all(c == 0 for c in cap), f"cap dư: {[c for c in cap if c]}"

    # ---- Kiểm tra ----
    use = Counter()
    name_dup_in_shop = 0
    sets = []
    for s in shops:
        assert len(s) == PER, f"shop size {len(s)} != {PER}"
        items = [r[ITEM_COL] for r in s]
        assert len(set(items)) == PER, "có listing lặp trong shop!"
        names = [norm_name(r[NAME_COL]) for r in s]
        name_dup_in_shop += sum(v - 1 for v in Counter(names).values() if v > 1)
        for r in s:
            use[r[ITEM_COL]] += 1
        sets.append(set(items))
    uc = Counter(use.values())
    # overlap giữa các cặp shop (mẫu thống kê)
    ov_min, ov_max, ov_sum, npair = 10**9, 0, 0, 0
    for i in range(N_SHOPS):
        for j in range(i + 1, N_SHOPS):
            o = len(sets[i] & sets[j])
            ov_min = min(ov_min, o); ov_max = max(ov_max, o); ov_sum += o; npair += 1

    print("\n--- KIỂM TRA ---")
    print("Mỗi shop:", PER, "sp khác nhau  -> OK")
    print("Listing lặp trong 1 shop:", 0, " (đảm bảo bằng thuật toán)")
    print("Cặp listing TRÙNG TÊN trong 1 shop:", name_dup_in_shop)
    print("Số lần dùng mỗi sp:", dict(sorted(uc.items())))
    print(f"Overlap giữa 2 shop: min={ov_min} avg={ov_sum/npair:.0f} max={ov_max} (trên {PER})")

    # ---- Ghi file ----
    out = Workbook(write_only=True)
    for i in range(1, N_SHOPS + 1):
        ws = out.create_sheet(title=f"shop{i}")
        rows = shops[i - 1]
        rows.sort(key=lambda r: (str(r[CAT_COL] or ""), norm_name(r[NAME_COL])))
        ws.append(header)
        for r in rows:
            ws.append(r)
    out.save(OUT)
    print("\nĐã lưu:", OUT)

if __name__ == "__main__":
    main()
